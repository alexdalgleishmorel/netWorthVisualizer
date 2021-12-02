# https://blog.quantinsti.com/stock-market-data-analysis-python/ for the yfinance and plotting
# information, this is what's used to visualize the data

import sys
import pandas as pd
import yfinance
import matplotlib.pyplot as plt
from datetime import date
import math
import userInterface as interface
import globals as globals
from http.server import BaseHTTPRequestHandler, HTTPServer
import time
import openpyxl
import pathlib
import speculation

hostName = "localhost"
serverPort = 8080

HOLDINGS = r"\holdings.xlsx"
DATALOG = r"\dataLog.xlsx"
EXCEL_FILE_LOCATION = str(pathlib.Path(__file__).parent.resolve())+HOLDINGS
DATALOG_LOCATION = str(pathlib.Path(__file__).parent.resolve())+DATALOG

# This is a class that creates asset objects, which can be a cryptocurrency or stock specified
# by the user
class Asset:

    # This method is executed each time a new Asset object is created
    def __init__(self, assetName, ticker, currency, dataframe, sharesArray, priceArray, datesArray):
        self.name = assetName
        self.ticker = ticker
        self.currency = currency

        if self.ticker != "Bank" and self.ticker != "cadCASH":
            self.currentPrice = get_current_price(ticker)
        else:
            self.currentPrice = None

        # If the user has the asset in excel, then this segment of the code will execute to calculate
        # certain metrics concerning their asset and initialize asset attributes with the correct values
        if not dataframe.empty and self.ticker != "Bank" and self.ticker != "cadCASH":
            self.userHolds = True
            self.sharesList = sharesArray
            self.priceList = priceArray
            self.datesList = datesArray

            self.buyMagnitudes, self.sellMagnitudes = createMagnitudes(dataframe)
            
            # Summing together all the purchased shares
            count = 0
            shares = 0
            for share in sharesArray:
                if math.isnan(share):
                    break
                shares += share
                count += 1
                count = 0
            
            self.sellPrices = []
            self.sellDates = []
            # Subtracting from the total shares the amount of shares that have been sold
            for share in dataframe['Shares Sold']:
                if math.isnan(share):
                    break
                # Adding the sell prices and dates as attributes of asset to be graphed later
                self.sellPrices.append(dataframe['Sell Price'][count])
                self.sellDates.append(dataframe['Sell Date'][count])
                shares -= share
                count += 1
            
            self.shares = shares
            # Calling the calculate book value method and associating the returned value with asset's book value
            if self.shares == 0:
                return
            self.bookValue = calculateNetBookValue(dataframe)
            self.averageCost = self.bookValue/self.shares
            self.marketValue = shares*self.currentPrice
            self.gainLossD = (self.marketValue)-(self.shares*self.averageCost)
            self.gainLossP = (self.gainLossD/(self.shares*self.averageCost))*100
        
        else:
            # User doesn't have any holdings of this currency
            self.userHolds = False
            self.sharesList = None
            self.priceList = None
            self.datesList = None
            self.shares = None
            self.bookValue = None
            self.averageCost = None
            self.marketValue = None

    def showHistory(self):
        if self.name == "Bank":
            plotNetWorth()
            return
        plotAdjClose(self.name)


def calculateNetBookValue(dataframe):
    sellDateCount, sellPriceCount, bookValue = 0, 0, 0
    lastBuy = -1
    for sale in dataframe['Shares Sold']:
        if math.isnan(sale):
            break
        # Get the sell date to cross reference for the buys
        sellDate = dataframe['Sell Date'][sellDateCount]
        sellPrice = dataframe['Sell Price'][sellPriceCount]
        # Sum all the buys up to the sell date
        sum = 0
        buyDateCount = 0
        for buy in dataframe['Shares']:
            if math.isnan(buy):
                break
            elif dataframe['Date'][buyDateCount] >= sellDate:
                break
            elif buyDateCount <= lastBuy:
                buyDateCount += 1
                continue
            else:
                sum += buy
                buyDateCount += 1
                lastBuy += 1
        sum = sum*sellPrice - (sale*sellPrice)
        bookValue += sum
        sellDateCount += 1
        sellPriceCount += 1

    lastBuy+=1
    counter = 0
    for buy in dataframe['Shares']:
        if math.isnan(buy):
            break
        if counter < lastBuy:
            counter+=1
            continue
        else:
            sum = buy*dataframe['Price'][counter]
            bookValue += sum
            counter+=1
    
    return bookValue
        

def get_current_price(symbol):
    ticker = yfinance.Ticker(symbol)
    #todays_data = ticker.history(period='1d')
    #return todays_data['Close'][0]
    data = ticker.history()
    return (data.tail(1)['Close'].iloc[0])

def processPurchaseData(dataframe):
    sharesArray = []
    for shareValues in dataframe['Shares']:
        if math.isnan(shareValues):
            break
        sharesArray.append(shareValues)
        
    priceArray = []
    for priceValues in dataframe['Price']:
        if math.isnan(priceValues):
            break
        priceArray.append(priceValues)
        
    datesArray = []
    for dateValues in dataframe['Date']:
        if pd.isnull(dateValues):
            break
        datesArray.append(dateValues)

    return sharesArray, priceArray, datesArray


def createMagnitudes(dataframe):
    buyMagnitudes = []
    sellMagnitudes = []

    count = 0
    for buy in dataframe['Shares']:
        if math.isnan(buy):
            break
        buyMagnitudes.append((buy*dataframe['Price'][count])/10)
        count += 1
    
    count = 0
    for sell in dataframe['Shares Sold']:
        if math.isnan(sell):
            break
        sellMagnitudes.append((sell*dataframe['Sell Price'][count])/10)
        count += 1

    return buyMagnitudes, sellMagnitudes

def plotAdjClose(assetName):
    # First we must find the right asset within the user's asset list
    assetToPlot = None
    for asset in globals.assetListCAD:
        if asset.name == assetName:
            assetToPlot = asset
            break
    if assetToPlot == None:
        for asset in globals.assetListUSD:
            if asset.name == assetName:
                assetToPlot = asset
                break


    todays_date = date.today()
    data = yfinance.download(assetToPlot.ticker, '2015-01-01', todays_date)
    
    # Plot adjusted close price data
    data['Adj Close'].plot(label = "{0} Price".format(assetToPlot.ticker))

    plt.axhline(y=assetToPlot.averageCost, color='cyan', label='Average Cost')

    buyPlotPoints = []
    sellPlotPoints = []
    count = 0
    # Creating list of buy and sell prices that will be consistent with the plotted line
    for buyPointDate in assetToPlot.datesList:
        buyPlotPoints.append(data['Adj Close'][buyPointDate])
        count += 1
    count = 0
    for sellPointDate in assetToPlot.sellDates:
        sellPlotPoints.append(data['Adj Close'][sellPointDate])
        count += 1

    if assetToPlot.userHolds:
        # Show purchase points on graph
        plt.scatter(x=assetToPlot.datesList, y=buyPlotPoints, s=assetToPlot.buyMagnitudes, alpha=0.5, marker='o', color="green", label = "Purchases")
        #plt.scatter(x=assetToPlot.datesList, y=assetToPlot.priceList, s=assetToPlot.buyMagnitudes, alpha=0.5, marker='o', color="green", label = "Purchases")
        # Show sell points on graph
        plt.scatter(x=assetToPlot.sellDates, y=sellPlotPoints, s=assetToPlot.sellMagnitudes, alpha=0.5, marker='o', color="red", label = "Sales")
        #plt.scatter(x=assetToPlot.sellDates, y=assetToPlot.sellPrices, s=assetToPlot.sellMagnitudes, alpha=0.5, marker='o', color="red", label = "Sales")

    plt.legend()

    # Define the label for the title of the figure
    plt.title("Adjusted Close Price of {0}".format(assetToPlot.name), fontsize=16)
    
    # Define the labels for x-axis and y-axis
    plt.ylabel('Price', fontsize=14)
    plt.xlabel('Year', fontsize=14)
    
    # Plot the grid lines
    plt.grid(which="major", color='k', linestyle='-.', linewidth=0.5)
    
    # Show the plot
    plt.show()

def plotNetWorth():
    dataframe = pd.read_excel (r'{0}'.format(DATALOG_LOCATION), sheet_name = "Financial History")

    plt.plot(dataframe['Date'], dataframe['Book'], label="Book Value")
    plt.plot(dataframe['Date'], dataframe['Net Wrth'], label="Net Worth")

    # Define the label for the title of the figure
    plt.title("Net Worth vs Book Value", fontsize=16)
    
    # Define the labels for x-axis and y-axis
    plt.ylabel('Value in $CAD', fontsize=14)
    plt.xlabel('Time', fontsize=14)
    
    # Plot the grid lines
    plt.grid(which="major", color='k', linestyle='-.', linewidth=0.5)

    plt.legend()

    plt.show()


# This function asks the user for a ticker of a stock to add to their list of assets and then generates
# an asset object (see Asset class), including user holdings info if it exists on the Excel sheet
def createAssets():
    try:
        dataframe = pd.read_excel (r'{0}'.format(EXCEL_FILE_LOCATION), sheet_name = None)
    except:
        print("You don't have any purchase history within an excel file")
        dataframe = pd.DataFrame({'A' : []})
        globals.prospectList.append(Asset(ticker, ticker, "later", dataframe, None, None, None))
        return

    for sheet in dataframe:
        ticker = sheet
        if ticker == 'Bank':
            bank = Asset(ticker, ticker, "CAD", dataframe["Bank"], None, None, None)
            cashCAD = Asset("cadCASH", "cadCASH", "CAD", dataframe["Bank"], None, None, None)
            globals.assetListCAD.append(bank)
            globals.assetListCAD.append(cashCAD)
            bank.marketValue = dataframe["Bank"]['Chequing'][0] + dataframe["Bank"]['Savings'][0]
            cashCAD.marketValue = dataframe["Bank"]['TFSA CAD CASH'][0] + dataframe["Bank"]['TFSA USD CASH'][0]*get_current_price("CAD=X")
            bank.bookValue = bank.marketValue
            cashCAD.bookValue = cashCAD.marketValue
            globals.totalCash = bank.bookValue + cashCAD.bookValue
            continue

        assetDataframe = dataframe[ticker]
        sharesArray, priceArray, datesArray = processPurchaseData(assetDataframe)
        if assetDataframe['Convert from USD?'][0] == 'yes':
            asset = Asset(ticker, ticker, "USD", assetDataframe, sharesArray, priceArray, datesArray)
            if asset.shares != 0:
                globals.assetListUSD.append(Asset(ticker, ticker, "USD", assetDataframe, sharesArray, priceArray, datesArray))
        else:
            asset = Asset(ticker, ticker, "CAD", assetDataframe, sharesArray, priceArray, datesArray)
            if asset.shares != 0:
                globals.assetListCAD.append(Asset(ticker, ticker, "CAD", assetDataframe, sharesArray, priceArray, datesArray))

# This is the main function that executes the overall functionality of the program
def main():
    # Creating a globally accessible list that will contain all the stocks the user has declared
    # to have positions in or wants to track
    globals.initialize()

    createAssets()
    for asset in globals.assetListCAD:
        globals.netWorth += asset.marketValue
        globals.bookValue += asset.bookValue
        print("{0}: {1} shares".format(asset.name, asset.shares))

    for asset in globals.assetListUSD:
        globals.netWorth += asset.marketValue*get_current_price("CAD=X")
        globals.bookValue += asset.bookValue*get_current_price("CAD=X")
        print("{0}: {1} shares".format(asset.name, asset.shares))

    # Calculate net gain/loss
    globals.gainLossD = (globals.netWorth)-(globals.bookValue)
    globals.gainLossP = (globals.gainLossD/(globals.bookValue))*100

    # Log landmark of bookvalue, networth and date into dataLog excel sheet
    dataLogWorkbook = openpyxl.load_workbook(DATALOG_LOCATION)

    dataLog_sheet = dataLogWorkbook.active

    rowToWrite = 2
    while (dataLog_sheet.cell(row=rowToWrite, column=1).value != None):
        rowToWrite += 1
    
    if (dataLog_sheet.cell(row=rowToWrite-1, column=1).value.date() != date.today()):
        print("Logging daily history...")
        dataLog_sheet.cell(row=rowToWrite, column=1).value = date.today()
        dataLog_sheet.cell(row=rowToWrite, column=2).value = globals.bookValue
        dataLog_sheet.cell(row=rowToWrite, column=3).value = globals.netWorth
        
        dataLogWorkbook.save(DATALOG_LOCATION)

    interface.createMenu(True)

    print(pathlib.Path(__file__).parent.resolve())

    """
    
    if __name__ == "__main__":        
        webServer = HTTPServer((hostName, serverPort), MyServer)
        print("Server started http://%s:%s" % (hostName, serverPort))
        
        try:
            webServer.serve_forever()
        except KeyboardInterrupt:
            pass
        
        webServer.server_close()
        print("Server stopped.")

        """


class MyServer(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-type", "text/html")
        self.end_headers()
        self.wfile.write(bytes("<html><head><title>https://pythonbasics.org</title></head>", "utf-8"))
        self.wfile.write(bytes("<p>Request: %s</p>" % self.path, "utf-8"))
        self.wfile.write(bytes("<body>", "utf-8"))

        # CREATE METHOD TO UPDATE NETWORTH AND CALL IT HERE

        self.wfile.write(bytes("<p>Overall Net Worth: ${0:.2f} CAD</p>".format(globals.netWorth), "utf-8"))
        self.wfile.write(bytes("</body></html>", "utf-8"))

main()
